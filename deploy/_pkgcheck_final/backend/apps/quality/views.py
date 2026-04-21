import logging
import requests
from django.conf import settings
from django.db import transaction
from rest_framework import viewsets, status
from rest_framework.decorators import api_view, action
from rest_framework.response import Response
from rest_framework.permissions import AllowAny
from openpyxl import load_workbook
from .models import Brand, MoistureData
from .serializers import BrandSerializer, MoistureDataSerializer

logger = logging.getLogger(__name__)


class BrandViewSet(viewsets.ModelViewSet):
    queryset = Brand.objects.all()
    serializer_class = BrandSerializer
    permission_classes = [AllowAny]


class MoistureDataViewSet(viewsets.ModelViewSet):
    queryset = MoistureData.objects.select_related('brand').all()
    serializer_class = MoistureDataSerializer
    filterset_fields = ['brand']
    permission_classes = [AllowAny]

    @action(detail=False, methods=['post'])
    def import_excel(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': '请上传文件'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = load_workbook(file, data_only=True)
            result = {'created': 0, 'skipped': 0, 'errors': []}

            with transaction.atomic():
                for sheet in wb.worksheets:
                    brand_name = sheet.title.replace('（', '(').replace('）', ')')
                    brand, _ = Brand.objects.get_or_create(name=brand_name)

                    headers = {}
                    for col_idx, cell in enumerate(sheet[4], start=1):
                        if cell.value:
                            headers[col_idx] = str(cell.value).strip()

                    for row_idx, row in enumerate(sheet.iter_rows(min_row=5, values_only=True), start=5):
                        try:
                            data = {headers.get(i + 1): v for i, v in enumerate(row) if headers.get(i + 1)}
                            row_list = list(row)

                            sample_number = data.get('样品编号')
                            if not sample_number:
                                continue

                            if MoistureData.objects.filter(brand=brand, sample_number=sample_number).exists():
                                result['skipped'] += 1
                                continue

                            def clean_value(val):
                                if val is None or val == '' or val == '/':
                                    return None
                                return val

                            def clean_decimal(val):
                                val = clean_value(val)
                                if val is None:
                                    return None
                                try:
                                    return float(val)
                                except:
                                    return None

                            sampling_date = clean_value(data.get('取样日期'))
                            if sampling_date and hasattr(sampling_date, 'date'):
                                sampling_date = sampling_date.date()
                            else:
                                sampling_date = None

                            MoistureData.objects.create(
                                brand=brand,
                                sampling_date=sampling_date,
                                sample_number=str(sample_number),
                                machine_number=str(clean_value(data.get('机台号')) or ''),
                                machine_stage=str(clean_value(data.get('机台')) or ''),
                                finished_moisture=clean_decimal(row_list[6] if len(row_list) > 6 else None),
                                powder_rate=clean_decimal(data.get('含末率')),
                                addition_method=str(clean_value(data.get('加丝方式/加丝机')) or ''),
                                batch_number=str(clean_value(data.get('批次号')) or ''),
                                shift=str(clean_value(data.get('班次')) or ''),
                                drying_moisture=clean_decimal(data.get('烘丝水分')),
                                mixed_moisture=clean_decimal(data.get('混合丝水分')),
                                cabinet_moisture=clean_decimal(data.get('出柜水分')),
                                rolling_moisture=clean_decimal(row_list[15] if len(row_list) > 15 else None)
                            )
                            result['created'] += 1
                        except Exception as e:
                            logger.error(f'导入第{row_idx}行失败: {e}', exc_info=True)
                            result['errors'].append(f'第{row_idx}行: {str(e)}')

            return Response(result)
        except Exception as e:
            logger.error(f'导入Excel失败: {e}', exc_info=True)
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def search_knowledge(request):
    """
    查询 RAGflow 质量知识库
    请求体：{"question": "问题内容", "top_n": 5}
    """
    question = request.data.get('question', '').strip()
    if not question:
        return Response({'error': '请输入查询内容'}, status=400)

    top_n = int(request.data.get('top_n', 5))

    if not settings.RAGFLOW_API_KEY or not settings.RAGFLOW_BASE_URL:
        return Response({'error': 'RAGflow 未配置，请联系管理员'}, status=503)

    dataset_id = settings.RAGFLOW_QUALITY_DATASET_ID
    if not dataset_id:
        return Response({'error': '质量知识库未配置'}, status=503)

    url = f"{settings.RAGFLOW_BASE_URL}/api/v1/retrieval"
    headers = {
        "Authorization": f"Bearer {settings.RAGFLOW_API_KEY}",
        "Content-Type": "application/json",
    }
    payload = {
        "question": question,
        "dataset_ids": [dataset_id],
        "top_n": top_n,
    }

    try:
        resp = requests.post(url, json=payload, headers=headers, timeout=15)
        resp.raise_for_status()
        data = resp.json()
        chunks = data.get('data', {}).get('chunks', [])

        results = [
            {
                'content': c.get('content', ''),
                'document_name': c.get('document_keyword', '') or c.get('docnm_kwd', ''),
                'score': round(c.get('similarity', 0), 3),
            }
            for c in chunks
        ]

        logger.info(f"质量知识库查询: '{question}' 返回 {len(results)} 条")
        return Response({'question': question, 'results': results, 'total': len(results)})

    except requests.exceptions.ConnectionError:
        return Response({'error': '无法连接到 RAGflow 服务'}, status=503)
    except Exception as e:
        logger.error(f"质量知识库查询失败: {e}")
        return Response({'error': f'查询失败: {str(e)}'}, status=500)
