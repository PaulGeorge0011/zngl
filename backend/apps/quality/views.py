import logging

import requests
from django.conf import settings
from django.db import transaction
from openpyxl import load_workbook
from rest_framework import status, viewsets
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.permissions import AllowAny
from rest_framework.response import Response

from .analysis_service import build_daily_report, build_quality_summary, detect_moisture_anomalies
from .models import Brand, MoistureData
from .serializers import BrandSerializer, MoistureDataSerializer

logger = logging.getLogger(__name__)


class BrandViewSet(viewsets.ModelViewSet):
    queryset = Brand.objects.all()
    serializer_class = BrandSerializer
    permission_classes = [AllowAny]


class MoistureDataViewSet(viewsets.ModelViewSet):
    queryset = MoistureData.objects.select_related('brand').all()
    serializer_class = MoistureDataSerializer
    filterset_fields = ['brand']
    permission_classes = [AllowAny]

    @action(detail=False, methods=['post'])
    def import_excel(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': '请上传文件'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            workbook = load_workbook(file, data_only=True)
            result = {'created': 0, 'skipped': 0, 'errors': []}

            with transaction.atomic():
                for sheet in workbook.worksheets:
                    brand_name = sheet.title.replace('（', '(').replace('）', ')')
                    brand, _ = Brand.objects.get_or_create(name=brand_name)

                    headers = {}
                    for col_idx, cell in enumerate(sheet[4], start=1):
                        if cell.value:
                            headers[col_idx] = str(cell.value).strip()

                    for row_idx, row in enumerate(sheet.iter_rows(min_row=5, values_only=True), start=5):
                        try:
                            data = {headers.get(i + 1): value for i, value in enumerate(row) if headers.get(i + 1)}
                            row_list = list(row)
                            sample_number = data.get('样品编号')
                            if not sample_number:
                                continue

                            if MoistureData.objects.filter(brand=brand, sample_number=sample_number).exists():
                                result['skipped'] += 1
                                continue

                            def clean_value(value):
                                if value in (None, '', '/'):
                                    return None
                                return value

                            def clean_decimal(value):
                                value = clean_value(value)
                                if value is None:
                                    return None
                                try:
                                    return float(value)
                                except (TypeError, ValueError):
                                    return None

                            sampling_date = clean_value(data.get('取样日期'))
                            if sampling_date and hasattr(sampling_date, 'date'):
                                sampling_date = sampling_date.date()
                            else:
                                sampling_date = None

                            MoistureData.objects.create(
                                brand=brand,
                                sampling_date=sampling_date,
                                sample_number=str(sample_number),
                                machine_number=str(clean_value(data.get('机台号')) or ''),
                                machine_stage=str(clean_value(data.get('机台')) or ''),
                                finished_moisture=clean_decimal(row_list[6] if len(row_list) > 6 else None),
                                powder_rate=clean_decimal(data.get('含末率')),
                                addition_method=str(clean_value(data.get('加丝方式/加丝机')) or ''),
                                batch_number=str(clean_value(data.get('批次号')) or ''),
                                shift=str(clean_value(data.get('班次')) or ''),
                                drying_moisture=clean_decimal(data.get('烘丝水分')),
                                mixed_moisture=clean_decimal(data.get('混合丝水分')),
                                cabinet_moisture=clean_decimal(data.get('出柜水分')),
                                rolling_moisture=clean_decimal(row_list[15] if len(row_list) > 15 else None),
                            )
                            result['created'] += 1
                        except Exception as exc:
                            logger.error('导入第 %s 行失败: %s', row_idx, exc, exc_info=True)
                            result['errors'].append(f'第 {row_idx} 行: {exc}')

            return Response(result)
        except Exception as exc:
            logger.error('导入 Excel 失败: %s', exc, exc_info=True)
            return Response({'error': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([AllowAny])
def search_knowledge(request):
    question = request.data.get('question', '').strip()
    if not question:
        return Response({'error': '请输入查询内容'}, status=400)

    top_n = int(request.data.get('top_n', 5))
    if not settings.RAGFLOW_API_KEY or not settings.RAGFLOW_BASE_URL:
        return Response({'error': 'RAGflow 未配置，请联系管理员'}, status=503)

    dataset_id = settings.RAGFLOW_QUALITY_DATASET_ID
    if not dataset_id:
        return Response({'error': '质量知识库未配置'}, status=503)

    url = f'{settings.RAGFLOW_BASE_URL}/api/v1/retrieval'
    headers = {
        'Authorization': f'Bearer {settings.RAGFLOW_API_KEY}',
        'Content-Type': 'application/json',
    }
    payload = {
        'question': question,
        'dataset_ids': [dataset_id],
        'top_n': top_n,
    }

    try:
        resp = requests.post(url, json=payload, headers=headers, timeout=15)
        resp.raise_for_status()
        data = resp.json()
        chunks = data.get('data', {}).get('chunks', [])
        results = [
            {
                'content': chunk.get('content', ''),
                'document_name': chunk.get('document_keyword', '') or chunk.get('docnm_kwd', ''),
                'score': round(chunk.get('similarity', 0), 3),
            }
            for chunk in chunks
        ]
        logger.info("质量知识库查询 '%s' 返回 %s 条", question, len(results))
        return Response({'question': question, 'results': results, 'total': len(results)})
    except requests.exceptions.ConnectionError:
        return Response({'error': '无法连接到 RAGflow 服务'}, status=503)
    except Exception as exc:
        logger.error('质量知识库查询失败: %s', exc)
        return Response({'error': f'查询失败: {exc}'}, status=500)


@api_view(['GET'])
@permission_classes([AllowAny])
def moisture_analysis_summary(request):
    brand_id = request.query_params.get('brand')
    brand_id = int(brand_id) if brand_id else None
    return Response(build_quality_summary(brand_id))


@api_view(['GET'])
@permission_classes([AllowAny])
def moisture_anomaly_alerts(request):
    brand_id = request.query_params.get('brand')
    limit = int(request.query_params.get('limit', 50))
    brand_id = int(brand_id) if brand_id else None
    return Response(detect_moisture_anomalies(brand_id=brand_id, limit=limit))


@api_view(['GET'])
@permission_classes([AllowAny])
def moisture_daily_report(request):
    brand_id = request.query_params.get('brand')
    report_date = request.query_params.get('date')
    brand_id = int(brand_id) if brand_id else None
    return Response(build_daily_report(report_date=report_date, brand_id=brand_id))
